import os
from openpyxl import load_workbook

# === 第一步：读取 Excel 文件第一列内容 ===
excel_path = 'standard.xlsx'  # Excel 文件名
wb = load_workbook(excel_path)
sheet = wb.active

# 读取第一列，去除空值和空白字符
lines_to_search = []
for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
    cell_value = row[0]
    if cell_value:
        lines_to_search.append(str(cell_value).strip())

# === 第二步：递归查找所有 .py 文件 ===
all_py_files = []
for root, dirs, files in os.walk(os.getcwd()):
    for file in files:
        if file.endswith('.py') and file != 'search_from_excel.py':
            full_path = os.path.join(root, file)
            all_py_files.append(full_path)

# === 第三步：搜索每一行内容是否出现在某些文件中 ===
for line in lines_to_search:
    print(f"\n🔍 正在查找内容：{line}")
    found = False
    for py_file in all_py_files:
        try:
            with open(py_file, 'r', encoding='utf-8') as f:
                content = f.read()
                if line in content:
                    print(f"✅ 在文件中找到匹配：{py_file}")
                    found = True
        except Exception as e:
            print(f"⚠️ 无法读取文件 {py_file}：{e}")
    if not found:
        print("❌ 没有找到匹配内容。")
